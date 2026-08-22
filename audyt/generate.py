#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generator artefaktów audytu i planu startowego Verris.

Źródłem prawdy są pliki CSV w `audyt/dane/`. Wszystko poniżej jest z nich wyliczane —
XLSX-y i dashboardy są WIDOKAMI. Ręczna edycja wyniku zniknie przy następnym uruchomieniu.

Użycie:
    python3 audyt/generate.py            # zbuduj wszystko
    python3 audyt/generate.py --sprawdz  # tylko walidacja spójności, bez zapisu
    python3 audyt/generate.py --status   # krótkie podsumowanie stanu

Wymaga: openpyxl.  Instalacja: pip install openpyxl
"""
import argparse, csv, json, datetime, pathlib, sys, re

BASE = pathlib.Path(__file__).resolve().parent
REPO = BASE.parent
DANE, SZAB = BASE / "dane", BASE / "szablony"
OUT_A = REPO / "audyt-parytetu-2026-08"
OUT_P = REPO / "plan-startowy-2026-08"

F = "Arial"
NAVY, RED, BLUE = "1F3A5F", "C00000", "1F6FB2"


# ───────────────────────────── wczytanie danych ─────────────────────────────
def wczytaj():
    def rows(name):
        with open(DANE / name, encoding="utf-8") as f:
            r = list(csv.reader(f))
        # Kontrola TUTAJ, a nie w sprawdz(), bo loader czyta pola po indeksie —
        # wiersz o złej liczbie kolumn przesuwa całą resztę i wstawia w plan
        # cudzą treść, zanim sprawdz() zdąży cokolwiek powiedzieć. Czasem
        # kończy się to wyjątkiem (gdy przesunięty tekst trafi w int()),
        # a czasem po cichu — i ten drugi przypadek jest gorszy.
        # Patrz PB-01 i „36,59 zł netto" bez cudzysłowów.
        if r:
            n = len(r[0])
            zle = [(nr, w) for nr, w in enumerate(r[1:], start=2) if len(w) != n]
            if zle:
                opis = "; ".join(
                    f"{name}:{nr} ma {len(w)} kolumn zamiast {n} (pole 1: {w[0] if w else '?'})"
                    for nr, w in zle
                )
                raise SystemExit(
                    f"BŁĄD DANYCH — {opis}. "
                    f"Najczęstsza przyczyna: przecinek w polu bez cudzysłowów."
                )
        return r[0], r[1:]

    head, macierz = rows("macierz.csv")
    _, sprinty = rows("sprinty.csv")
    _, pb = rows("zadania_pb.csv")
    _, epiki = rows("epiki.csv")
    _, fazy = rows("fazy.csv")
    cfg = json.load(open(DANE / "konfiguracja.json", encoding="utf-8"))
    return {
        "head": head,
        "macierz": macierz,
        "wg_id": {r[0]: r for r in macierz},
        "sprinty": {int(s[0]): {"cel": s[1], "audyt": s[2], "pb": s[3], "ryzyko": s[4]} for s in sprinty},
        "pb": {p[0]: {"tytul": p[1], "h": int(p[2]), "prio": p[3], "opis": p[4], "dod": p[5],
                       "zamkniete": (p[6] if len(p) > 6 else "")} for p in pb},
        "epiki": [{"id": e[0], "nazwa": e[1], "prio": e[2], "kw": e[3], "zakres": e[4], "why": e[5]} for e in epiki],
        "fazy": [{"od": int(f_[0]), "do": int(f_[1]), "tytul": f_[2], "opis": f_[3]} for f_ in fazy],
        "cfg": cfg,
    }


def pozycje_sprintu(D, n):
    """[(id, tytuł, godziny, typ, priorytet, dowód, definicja ukończenia, kontekst)]"""
    H = D["cfg"]["godziny_nakladu"]
    s, out = D["sprinty"][n], []
    for tok in filter(None, s["audyt"].split(";")):
        if ":" in tok:
            i, h = tok.split(":"); h = int(h); czesc = " (część)"
        else:
            i, h, czesc = tok, H[D["wg_id"][tok][12]], ""
        r = D["wg_id"][i]
        dod = D["cfg"]["definicje_ukonczenia_wg_stanu"].get(r[8], D["cfg"]["definicje_ukonczenia_wg_stanu"]["BRAK"])
        out.append((i, r[2] + czesc, h, "audyt", r[11], r[9], dod, r[13]))
    for i in filter(None, s["pb"].split(";")):
        p = D["pb"][i]
        out.append((i, p["tytul"], p["h"], "poza audytem", p["prio"], "—", p["dod"], p["opis"]))
    return out


def godziny(D, n):
    return sum(x[2] for x in pozycje_sprintu(D, n))


def daty(D, n):
    a = datetime.date.fromisoformat(D["cfg"]["start"]) + datetime.timedelta(weeks=n - 1)
    return a, a + datetime.timedelta(days=4)


def faza(D, n):
    for f_ in D["fazy"]:
        if f_["od"] <= n <= f_["do"]:
            return f_["tytul"]
    return ""


def epik_dla(D, r):
    return D["cfg"]["wyjatki_epik"].get(r[0]) or D["cfg"]["kat2epik"].get(r[1], "E-16")


def reszta(D):
    zapl = {x[0] for n in D["sprinty"] for x in pozycje_sprintu(D, n)}
    H = D["cfg"]["godziny_nakladu"]
    r = [x for x in D["macierz"] if x[10] in ("LUKA", "CZĘŚCIOWY") and x[0] not in zapl]
    return r, sum(H.get(x[12], 0) for x in r)


# ───────────────────────────── walidacja ─────────────────────────────
def sprawdz(D):
    bledy, ostrz = [], []
    ids = [x[0] for x in D["macierz"]]
    if len(ids) != len(set(ids)):
        bledy.append("zduplikowane ID w macierz.csv")

    dozwolone = {
        8: {"DZIAŁA", "CZĘŚCIOWE", "FLAGA", "ATRAPA", "ENDPOINT BEZ UI", "BRAK", "b.d."},
        10: {"PARYTET", "PRZEWAGA", "CZĘŚCIOWY", "LUKA", "POZA ZAKRESEM"},
        11: {"BLOKER STARTU", "WYSOKA", "ŚREDNIA", "NISKA", "—"},
        12: {"S", "M", "L", "—", ""},
    }
    for r in D["macierz"]:
        for kol, ok in dozwolone.items():
            if r[kol] not in ok:
                bledy.append(f"{r[0]}: niedozwolona wartość w kolumnie {kol}: {r[kol]!r}")

    przypisane = {x[0] for n in D["sprinty"] for x in pozycje_sprintu(D, n)}
    for r in D["macierz"]:
        if r[11] == "BLOKER STARTU" and r[0] not in przypisane:
            bledy.append(f"{r[0]} jest blokerem startu i nie ma sprintu")
    for tok in [t for s in D["sprinty"].values() for t in filter(None, s["audyt"].split(";"))]:
        i = tok.split(":")[0]
        if i not in D["wg_id"]:
            bledy.append(f"sprint odwołuje się do nieistniejącego ID: {i}")

    pbp = {i for s in D["sprinty"].values() for i in filter(None, s["pb"].split(";"))}
    for i in D["pb"]:
        if i not in pbp:
            bledy.append(f"zadanie {i} nie ma sprintu")

    cap = D["cfg"]["sprint_godzin"]
    for n in sorted(D["sprinty"]):
        h = godziny(D, n)
        if h > cap + 2:
            ostrz.append(f"sprint {n} przeciążony: {h} h przy pojemności {cap} h")

    for r in D["macierz"]:
        if r[10] == "POZA ZAKRESEM" and not r[13].strip():
            ostrz.append(f"{r[0]}: POZA ZAKRESEM bez uzasadnienia — to LUKA udająca decyzję")
        if r[8] == "DZIAŁA" and r[9].strip() in ("", "—"):
            ostrz.append(f"{r[0]}: stan DZIAŁA bez dowodu plik:linia")
    return bledy, ostrz


# ───────────────────────────── pomocnicze XLSX ─────────────────────────────
def _xl():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return openpyxl, Font, PatternFill, Alignment, Border, Side


def stylizuj(ws, szer, zawijaj, nrows, zamroz="C2"):
    _, Font, PatternFill, Alignment, Border, Side = _xl()
    from openpyxl.utils import get_column_letter
    thin = Side(style="thin", color="D0D0D0")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, x in enumerate(szer, 1):
        ws.column_dimensions[get_column_letter(i)].width = x
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(name=F, bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=nrows, max_col=len(szer)):
        for c in row:
            c.font = Font(name=F, size=10)
            c.border = bd
            c.alignment = Alignment(vertical="top", wrap_text=(c.column in zawijaj))
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = zamroz
    ws.auto_filter.ref = f"A1:{get_column_letter(len(szer))}{nrows}"


def regula(ws, kol, txt, bg, fg="000000"):
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill, Font
    ws.conditional_formatting.add(
        f"{kol}2:{kol}{ws.max_row}",
        CellIsRule(operator="equal", formula=[f'"{txt}"'],
                   fill=PatternFill("solid", fgColor=bg),
                   font=Font(name=F, size=10, bold=True, color=fg)))


# ───────────────────────────── 1. macierz XLSX ─────────────────────────────
def buduj_macierz_xlsx(D):
    openpyxl, Font, PatternFill, Alignment, _, _ = _xl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Macierz"
    ws.append(D["head"])
    for r in D["macierz"]:
        ws.append(r)
    stylizuj(ws, [8, 44, 13, 10, 10, 13, 11, 11, 17, 46, 13, 15, 8, 60], (2, 10, 14), ws.max_row)
    N = ws.max_row
    for kol in "DEFGH":
        for t, bg in [("TAK", "D6EFD8"), ("CZĘŚĆ", "FFF2CC"), ("PŁATNE", "FCE4D6"),
                      ("NIE", "F2F2F2"), ("b.d.", "EDEDED")]:
            regula(ws, kol, t, bg)
    for t, bg in [("DZIAŁA", "D6EFD8"), ("CZĘŚCIOWE", "FFF2CC"), ("FLAGA", "FFE699"),
                  ("ATRAPA", "F8CBAD"), ("ENDPOINT BEZ UI", "F8CBAD"), ("BRAK", "F4B6B6")]:
        regula(ws, "I", t, bg)
    for t, bg in [("PARYTET", "D6EFD8"), ("PRZEWAGA", "BDD7EE"), ("CZĘŚCIOWY", "FFF2CC"),
                  ("LUKA", "F4B6B6"), ("POZA ZAKRESEM", "EDEDED")]:
        regula(ws, "K", t, bg)
    regula(ws, "L", "BLOKER STARTU", RED, "FFFFFF")
    regula(ws, "L", "WYSOKA", "F8CBAD")
    regula(ws, "L", "ŚREDNIA", "FFF2CC")

    s = wb.create_sheet("Podsumowanie", 0)
    s.column_dimensions["A"].width = 36
    for c in "BCDEF":
        s.column_dimensions[c].width = 15

    def tyt(cell, txt, size=13):
        s[cell] = txt
        s[cell].font = Font(name=F, bold=True, size=size, color=NAVY)

    tyt("A1", "Verris — parytet funkcji wobec rynku", 15)
    s["A2"] = f"Wygenerowane {datetime.date.today()} z audyt/dane/macierz.csv · nie edytuj tego pliku ręcznie"
    s["A2"].font = Font(name=F, size=10, italic=True, color="606060")

    bloki = [
        (4, "Werdykt parytetu", "K", ["PARYTET", "PRZEWAGA", "CZĘŚCIOWY", "LUKA", "POZA ZAKRESEM"]),
        (13, "Stan implementacji", "I", ["DZIAŁA", "CZĘŚCIOWE", "FLAGA", "ATRAPA", "ENDPOINT BEZ UI", "BRAK", "b.d."]),
        (24, "Krytyczność luk", "L", ["BLOKER STARTU", "WYSOKA", "ŚREDNIA", "NISKA"]),
    ]
    for r0, tytul, kol, wart in bloki:
        tyt(f"A{r0}", tytul)
        s[f"A{r0+1}"], s[f"B{r0+1}"] = "Wartość", "Pozycji"
        for c in (f"A{r0+1}", f"B{r0+1}"):
            s[c].fill = PatternFill("solid", fgColor=NAVY)
            s[c].font = Font(name=F, bold=True, color="FFFFFF", size=10)
        for i, v in enumerate(wart):
            r = r0 + 2 + i
            s[f"A{r}"] = v
            s[f"B{r}"] = f'=COUNTIF(Macierz!${kol}$2:${kol}${N},A{r})'
            s[f"A{r}"].font = Font(name=F, size=10, bold=(v == "BLOKER STARTU"),
                                   color=(RED if v == "BLOKER STARTU" else "000000"))
            s[f"B{r}"].font = Font(name=F, size=10, bold=(v == "BLOKER STARTU"),
                                   color=(RED if v == "BLOKER STARTU" else "000000"))

    tyt("A31", "Pokrycie per kategoria")
    for i, h in enumerate(["Kategoria", "Pozycji", "Parytet+Przewaga", "Luki", "Blokery", "Pokrycie"]):
        c = s.cell(row=32, column=1 + i, value=h)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(name=F, bold=True, color="FFFFFF", size=10)
    for i, k in enumerate(sorted(D["cfg"]["kategorie"])):
        r = 33 + i
        s.cell(row=r, column=1, value=D["cfg"]["kategorie"][k]).font = Font(name=F, size=10)
        s.cell(row=r, column=2, value=f'=COUNTIF(Macierz!$B$2:$B${N},"{k}")')
        s.cell(row=r, column=3, value=f'=COUNTIFS(Macierz!$B$2:$B${N},"{k}",Macierz!$K$2:$K${N},"PARYTET")'
                                     f'+COUNTIFS(Macierz!$B$2:$B${N},"{k}",Macierz!$K$2:$K${N},"PRZEWAGA")')
        s.cell(row=r, column=4, value=f'=COUNTIFS(Macierz!$B$2:$B${N},"{k}",Macierz!$K$2:$K${N},"LUKA")')
        s.cell(row=r, column=5, value=f'=COUNTIFS(Macierz!$B$2:$B${N},"{k}",Macierz!$L$2:$L${N},"BLOKER STARTU")')
        cc = s.cell(row=r, column=6, value=f"=IFERROR(C{r}/B{r},0)")
        cc.number_format = "0%"
        for col in range(2, 7):
            s.cell(row=r, column=col).font = Font(name=F, size=10)

    r0 = 33 + len(D["cfg"]["kategorie"]) + 2
    s.cell(row=r0, column=1, value="UWAGA METODOLOGICZNA").font = Font(name=F, bold=True, size=11, color=RED)
    for i, t in enumerate([
        "Procent pokrycia jest metryką OPISOWĄ. Decyzję o starcie podejmuje się wyłącznie na podstawie liczby blokerów.",
        "Kolumny cPanel/Plesk/DirectAdmin pokazują możliwości silnika — nie to, co hosting faktycznie daje klientowi.",
        "Werdykt liczony jest wobec kolumny „Rynek PL”, nie wobec sumy możliwości paneli.",
        "Stan Verris ustalany z kodu (plik:linia), nie z dokumentacji projektu.",
        "„ATRAPA” = UI bez działającego backendu. „ENDPOINT BEZ UI” = backend gotowy, żaden panel go nie woła.",
        "Ten plik jest generowany. Zmiany wprowadzaj w audyt/dane/macierz.csv i uruchom audyt/generate.py.",
    ]):
        c = s.cell(row=r0 + 1 + i, column=1, value="• " + t)
        c.font = Font(name=F, size=10)

    OUT_A.mkdir(exist_ok=True)
    wb.save(OUT_A / "VERRIS_PARYTET_FUNKCJI_2026-08.xlsx")


# ───────────────────────────── 2. dashboard luk ─────────────────────────────
def buduj_dashboard_luk(D):
    idx = {n: i for i, n in enumerate(
        ["id", "kat", "f", "cp", "pl", "da", "rp", "nf", "st", "dow", "w", "kr", "nk", "u"])}
    rows = [{k: r[i] for k, i in idx.items()} for r in D["macierz"]]
    dane = json.dumps({"rows": rows, "kats": D["cfg"]["kategorie"]}, ensure_ascii=False)
    tpl = (SZAB / "dashboard_luki.html").read_text(encoding="utf-8")
    OUT_A.mkdir(exist_ok=True)
    (OUT_A / "VERRIS_LUKI_DASHBOARD.html").write_text(tpl.replace("__DATA__", dane), encoding="utf-8")


# ───────────────────────────── 3. backlog XLSX ─────────────────────────────
def zadania(D):
    out = []
    for n in sorted(D["sprinty"]):
        d0, d1 = daty(D, n)
        for i, t, h, typ, prio, dow, dod, ctx in pozycje_sprintu(D, n):
            r = D["wg_id"].get(i)
            out.append([i, t, typ,
                        (D["cfg"]["kategorie"].get(r[1], r[1]) if r else "Praca spoza audytu"),
                        prio, h, n, faza(D, n), str(d0), str(d1),
                        (r[8] if r else "—"), dow,
                        D["cfg"]["zaleznosci"].get(i, "—"), dod, ctx, "do zrobienia"])
    return out


def buduj_backlog_xlsx(D):
    openpyxl, Font, PatternFill, Alignment, _, _ = _xl()
    from openpyxl.formatting.rule import DataBarRule
    T = zadania(D)
    R, RH = reszta(D)
    OOS = [r for r in D["macierz"] if r[10] == "POZA ZAKRESEM"]
    EP = {e["id"]: e for e in D["epiki"]}
    cap = D["cfg"]["sprint_godzin"]
    H = D["cfg"]["godziny_nakladu"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Backlog do startu"
    ws.append(["ID", "Zadanie", "Typ", "Kategoria", "Priorytet", "h", "Sprint", "Faza", "Od", "Do",
               "Stan w kodzie", "Dowód (plik:linia)", "Zależy od", "Definicja ukończenia", "Kontekst", "Status"])
    for t in T:
        ws.append(t)
    stylizuj(ws, [8, 44, 13, 26, 17, 5, 7, 42, 11, 11, 17, 44, 40, 60, 60, 14], (2, 12, 13, 14, 15), ws.max_row)
    N = ws.max_row
    for t, bg, fg in [("BLOKER STARTU", RED, "FFFFFF"), ("BLOKER BIZNESOWY", RED, "FFFFFF"),
                      ("WYSOKA", "F8CBAD", "000000"), ("WYSOKI", "F8CBAD", "000000"),
                      ("ŚREDNIA", "FFF2CC", "000000"), ("ŚREDNI", "FFF2CC", "000000")]:
        regula(ws, "E", t, bg, fg)
    regula(ws, "P", "zrobione", "D6EFD8", "0C6B0C")
    regula(ws, "P", "w toku", "FFF2CC")

    sp = wb.create_sheet("Sprinty")
    sp.append(["Sprint", "Od", "Do", "Faza", "Cel sprintu", "Zadań", "Godzin", "Pojemność", "Obciążenie", "Ryzyko"])
    for n in sorted(D["sprinty"]):
        d0, d1 = daty(D, n)
        r = n + 1
        sp.append([n, str(d0), str(d1), faza(D, n), D["sprinty"][n]["cel"],
                   f"=COUNTIF('Backlog do startu'!$G$2:$G${N},A{r})",
                   f"=SUMIF('Backlog do startu'!$G$2:$G${N},A{r},'Backlog do startu'!$F$2:$F${N})",
                   cap, f"=IFERROR(G{r}/H{r},0)", D["sprinty"][n]["ryzyko"]])
    stylizuj(sp, [7, 11, 11, 40, 52, 7, 8, 11, 12, 92], (4, 5, 10), sp.max_row, "A2")
    for r in range(2, sp.max_row + 1):
        sp.cell(row=r, column=9).number_format = "0%"
    sp.conditional_formatting.add(f"G2:G{sp.max_row}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=cap, color="4A90D9"))

    ps = wb.create_sheet("Po starcie")
    ps.append(["ID", "Epik", "Nazwa epiku", "Kwartał", "Priorytet epiku", "Funkcja", "Kategoria",
               "Krytyczność", "h", "Stan w kodzie", "Dowód / uwagi"])
    for r in sorted(R, key=lambda r: (EP[epik_dla(D, r)]["kw"], epik_dla(D, r), r[0])):
        e = EP[epik_dla(D, r)]
        ps.append([r[0], e["id"], e["nazwa"], e["kw"], e["prio"], r[2],
                   D["cfg"]["kategorie"].get(r[1], r[1]), r[11], H.get(r[12], 0), r[8],
                   (r[9] if r[9] != "—" else r[13])])
    stylizuj(ps, [8, 8, 34, 11, 15, 46, 26, 15, 5, 17, 66], (3, 6, 11), ps.max_row, "A2")

    oz = wb.create_sheet("Poza zakresem")
    oz.append(["ID", "Funkcja", "Kategoria", "Uzasadnienie decyzji"])
    for r in OOS:
        oz.append([r[0], r[2], D["cfg"]["kategorie"].get(r[1], r[1]), r[13] or "—"])
    stylizuj(oz, [8, 46, 26, 88], (2, 4), oz.max_row, "A2")

    s = wb.create_sheet("Podsumowanie", 0)
    s.column_dimensions["A"].width = 40
    s.column_dimensions["B"].width = 78

    def tyt(cell, txt, size=13):
        s[cell] = txt
        s[cell].font = Font(name=F, bold=True, size=size, color=NAVY)

    tyt("A1", "Verris — backlog do startu", 15)
    s["A2"] = (f"{len(D['sprinty'])} sprintów · {cap} h netto na sprint · "
               f"wygenerowane {datetime.date.today()} z audyt/dane/")
    s["A2"].font = Font(name=F, size=10, italic=True, color="606060")
    tyt("A4", "Kluczowe liczby")
    for i, (k, v) in enumerate([
        ("Zadań w planie", f"=COUNTA('Backlog do startu'!$A$2:$A${N})"),
        ("Godzin w planie", f"=SUM('Backlog do startu'!$F$2:$F${N})"),
        ("Pojemność", len(D["sprinty"]) * cap),
        ("Zapas", "=B7-B6"),
        ("Blokerów", f"=COUNTIF('Backlog do startu'!$E$2:$E${N},\"BLOKER STARTU\")"
                     f"+COUNTIF('Backlog do startu'!$E$2:$E${N},\"BLOKER BIZNESOWY\")"),
        ("Zadań zrobionych", f"=COUNTIF('Backlog do startu'!$P$2:$P${N},\"zrobione\")"),
        ("Godzin zrobionych", f"=SUMIF('Backlog do startu'!$P$2:$P${N},\"zrobione\",'Backlog do startu'!$F$2:$F${N})"),
        ("Postęp", "=IFERROR(B12/B6,0)"),
    ]):
        r = 6 + i
        s[f"A{r}"], s[f"B{r}"] = k, v
        s[f"A{r}"].font = Font(name=F, size=10)
        s[f"B{r}"].font = Font(name=F, size=10, bold=True)
    s["B13"].number_format = "0%"
    s["A10"].font = Font(name=F, size=10, bold=True, color=RED)
    s["B10"].font = Font(name=F, size=11, bold=True, color=RED)

    tyt("A15", "Praca poza planem sprintów")
    for i, (k, v) in enumerate([("Pozycji w roadmapie po starcie", len(R)), ("Godzin", RH),
                                ("Sprintów po tyle godzin", round(RH / cap, 1)),
                                ("Pozycji poza zakresem", len(OOS))]):
        r = 16 + i
        s[f"A{r}"], s[f"B{r}"] = k, v
        s[f"A{r}"].font = Font(name=F, size=10)
        s[f"B{r}"].font = Font(name=F, size=10, bold=True)

    tyt("A21", "Jak używać tego pliku")
    for i, t in enumerate([
        "Edytuje się WYŁĄCZNIE kolumnę Status w arkuszu „Backlog do startu”: do zrobienia / w toku / zrobione.",
        "Liczby powyżej i obciążenie sprintów przeliczają się same.",
        "„Zrobione” = spełniona definicja ukończenia z kolumny N, nie „kod napisany”.",
        "Nic poniżej D2 (test w CI) nie jest zrobione. Pieniądze, dane i dostęp → D3. Backupy i DR → D4.",
        "Reszta pliku jest generowana z audyt/dane/ — zmiany rób tam i uruchom audyt/generate.py.",
    ]):
        s.cell(row=22 + i, column=1, value="• " + t).font = Font(name=F, size=10)

    OUT_P.mkdir(exist_ok=True)
    wb.save(OUT_P / "VERRIS_BACKLOG_STARTOWY.xlsx")


# ───────────────────────────── 4. plan MD ─────────────────────────────
def buduj_plan_md(D):
    L, w = [], None
    L = []
    w = L.append
    NS = len(D["sprinty"])
    cap = D["cfg"]["sprint_godzin"]
    total = sum(godziny(D, n) for n in D["sprinty"])
    R, RH = reszta(D)
    OOS = [r for r in D["macierz"] if r[10] == "POZA ZAKRESEM"]
    EP = {e["id"]: e for e in D["epiki"]}
    H = D["cfg"]["godziny_nakladu"]

    w("# Plan sprintów do startu — Verris\n")
    w(f"**Wygenerowany:** {datetime.date.today()} z `audyt/dane/` · **nie edytuj ręcznie**  ")
    w(f"**Podstawa:** audyt parytetu funkcji z 2026-08-20  ")
    w(f"**Pojemność:** 1 osoba, pełny etat, **{cap} h netto na sprint** · sprint = 1 tydzień  ")
    w(f"**Sprint 1:** {daty(D,1)[0]} · **Sprint {NS}:** {daty(D,NS)[0]}–{daty(D,NS)[1]}\n")
    w("---\n")
    w("## Liczba, od której trzeba zacząć\n")
    w(f"Domknięcie **wszystkich** luk z macierzy to **{total + RH} h** — przy {cap} h tygodniowo "
      f"około **{round((total+RH)/cap/4.3)} miesięcy pracy solo, bez jednego przychodu po drodze**. "
      "Taki plan nie jest planem startu, tylko sposobem, żeby nigdy nie wystartować.\n")
    w(f"Dlatego praca dzieli się na dwie części: **{NS} sprintów do startu** ({total} h) oraz "
      f"roadmapę po starcie ({RH} h, {len(R)} pozycji) rozpisaną na epiki kwartalne.\n")
    blk_bez_ksef = max(n for n in D["sprinty"]
                       for x in pozycje_sprintu(D, n)
                       if x[4] == "BLOKER STARTU" and x[0] not in ("M-14", "M-15", "M-16", "M-17"))
    w(f"- **{daty(D, blk_bez_ksef)[1]}** — koniec sprintu {blk_bez_ksef}, zamknięte wszystkie blokery **poza KSeF-em**.")
    w(f"- **{daty(D, NS)[1]}** — koniec sprintu {NS}, decyzja GO.\n")
    w("---\n")
    w("## Zasady obowiązujące w każdym sprincie\n")
    for i, t in enumerate([
        "**Każda naprawiona pozycja dostaje test, który najpierw czerwieni się na starym kodzie.** Test napisany po naprawie i od razu zielony nie dowodzi niczego.",
        "**Sprint kończy się, gdy definicja ukończenia jest spełniona, a nie gdy mija piątek.** Przesunięcie jest informacją; ukrycie przesunięcia jest porażką.",
        "**Status wg skali dowodu.** Nic poniżej D2 nie jest „zrobione”. Pieniądze, dane klienta i dostęp → D3. Backupy i DR → D4.",
        "**Zakaz formuły „warunkowe GO”.**",
        "**Nowa praca odkryta w sprincie nie wchodzi do niego** — trafia do backlogu. Wyjątek: bloker znaleziony przy naprawie innego blokera.",
        "**Każde zadanie ma plik w `docs/zadania/`**, każdy sprint podsumowanie w `docs/sprinty/`. Z tego składa się dokumentacja techniczna.",
        "**Każdy sprint kończy się aktualizacją `audyt/dane/macierz.csv`** i przebudową widoków. Procedura: `plan-startowy-2026-08/AKTUALIZACJA_AUDYTU.md`.",
    ], 1):
        w(f"{i}. {t}")
    w("\n---\n")

    for f_ in D["fazy"]:
        fh = sum(godziny(D, n) for n in range(f_["od"], f_["do"] + 1))
        w(f"# {f_['tytul']}\n")
        w(f"*Sprinty {f_['od']}–{f_['do']} · {fh} h · {daty(D,f_['od'])[0]} – {daty(D,f_['do'])[1]}*\n")
        w(f_["opis"] + "\n")
        for n in range(f_["od"], f_["do"] + 1):
            d0, d1 = daty(D, n)
            its = pozycje_sprintu(D, n)
            w(f"## Sprint {n} — {D['sprinty'][n]['cel']}\n")
            w(f"`{d0} – {d1}` · **{godziny(D,n)} h** z {cap} h pojemności\n")
            w("| ID | Zadanie | h | Priorytet | Dowód / kontekst |")
            w("|---|---|---|---|---|")
            for i, t, h, typ, prio, dow, dod, ctx in its:
                w(f"| `{i}` | {t} | {h} | {prio} | {(dow if dow != '—' else ctx)[:150]} |")
            w("")
            w("**Definicja ukończenia**\n")
            for i, t, h, typ, prio, dow, dod, ctx in its:
                w(f"- `{i}` — {dod}")
            w("- **Cały sprint** — `docs/zadania/` uzupełnione dla każdej pozycji, `docs/sprinty/SPRINT-%02d.md` napisane, "
              "`audyt/dane/macierz.csv` zaktualizowana, widoki przebudowane." % n)
            w("")
            w(f"**Ryzyko sprintu.** {D['sprinty'][n]['ryzyko']}\n")
        w("---\n")

    w("# Po starcie — roadmapa kwartalna\n")
    w(f"{len(R)} pozycji, {RH} h. Epiki, nie sprinty — kolejność zweryfikujemy danymi od pierwszych klientów.\n")
    eh, en = {}, {}
    for r in R:
        e = epik_dla(D, r)
        eh[e] = eh.get(e, 0) + H.get(r[12], 0)
        en[e] = en.get(e, 0) + 1
    w("| ID | Epik | Priorytet | Kwartał | Pozycji | h | Dlaczego teraz, a nie wcześniej |")
    w("|---|---|---|---|---|---|---|")
    for e in sorted(D["epiki"], key=lambda x: (x["kw"], x["id"])):
        w(f"| `{e['id']}` | {e['nazwa']} | {e['prio']} | {e['kw']} | {en.get(e['id'],0)} | {eh.get(e['id'],0)} | {e['why']} |")
    w("")
    for e in D["epiki"]:
        w(f"- **{e['id']} {e['nazwa']}** ({eh.get(e['id'],0)} h) — {e['zakres']}")
    w("\n---\n")
    w("# Czego świadomie nie robimy\n")
    w(f"{len(OOS)} pozycji ma werdykt POZA ZAKRESEM. To decyzje, nie przeoczenia — dlatego są wypisane. "
      "Jeżeli któraś wróci jako żądanie klienta, wraca też decyzja do przeglądu.\n")
    w("| ID | Funkcja | Uzasadnienie |")
    w("|---|---|---|")
    for r in OOS:
        w(f"| `{r[0]}` | {r[2]} | {r[13] or '—'} |")
    w("")
    w("Cel „100% pokrycia we wszystkich kategoriach” nie kształtuje tego planu. Punktem odniesienia jest mediana "
      "rynku PL plus to, co klient uznaje za standard — nie suma możliwości cPanela, Pleska i DirectAdmina. "
      "Pokrycie rośnie tu jako skutek uboczny zamykania rzeczy, które mają znaczenie.\n")
    OUT_P.mkdir(exist_ok=True)
    (OUT_P / "PLAN_SPRINTOW_2026-08.md").write_text("\n".join(L), encoding="utf-8")


# ───────────────────────────── 5. dashboard planu ─────────────────────────────

# ───────────────────────────── postęp ─────────────────────────────
#
# Liczony z macierzy i z zadania_pb.csv, nie z osobnej listy „co zrobione".
# Osobna lista rozjechałaby się z macierzą przy pierwszym pominiętym wpisie,
# a zielona linia pokazywałaby wtedy postęp, którego nie ma.

ZROBIONE_WERDYKTY = ("PARYTET", "PRZEWAGA")


def _zrobione(D, i, typ):
    """Czy pozycja jest zamknięta. Słownik audytu, nie moja ocena."""
    if typ == "audyt":
        r = D["wg_id"].get(i)
        return bool(r) and r[10] in ZROBIONE_WERDYKTY
    return bool(D["pb"].get(i, {}).get("zamkniete"))


def _data_zamkniecia(D, i, typ):
    if typ == "poza audytem":
        return D["pb"].get(i, {}).get("zamkniete") or None
    r = D["wg_id"].get(i)
    if not r:
        return None
    m = re.search(r"Zamkni[eę]te (\d{4}-\d{2}-\d{2})", r[13])
    return m.group(1) if m else None


def postep(D, dzis=None):
    dzis = dzis or datetime.date.today()
    start = datetime.date.fromisoformat(D["cfg"]["start"])
    cap = D["cfg"]["sprint_godzin"]

    sprinty = []
    for n in sorted(D["sprinty"]):
        poz = pozycje_sprintu(D, n)
        zrob = [x for x in poz if _zrobione(D, x[0], x[3])]
        d0, d1 = daty(D, n)
        sprinty.append({
            "n": n,
            "cel": D["sprinty"][n]["cel"],
            "od": str(d0), "do": str(d1),
            "pozycje": len(poz),
            "zrobione": len(zrob),
            "godziny": sum(x[2] for x in poz),
            "godzinyZrobione": sum(x[2] for x in zrob),
            "otwarte": [
                {"id": x[0], "tytul": x[1], "h": x[2],
                 "bloker": x[4] == "BLOKER STARTU" or "BLOKER" in str(x[4])}
                for x in poz if not _zrobione(D, x[0], x[3])
            ],
            "domkniete": [
                {"id": x[0], "tytul": x[1], "h": x[2], "data": _data_zamkniecia(D, x[0], x[3])}
                for x in zrob
            ],
        })

    godz_razem = sum(s["godziny"] for s in sprinty)
    godz_zrob = sum(s["godzinyZrobione"] for s in sprinty)
    poz_razem = sum(s["pozycje"] for s in sprinty)
    poz_zrob = sum(s["zrobione"] for s in sprinty)

    # Ile pełnych tygodni planu upłynęło. Ujemne przed startem planu —
    # i tak ma być, bo praca ruszyła wcześniej niż harmonogram.
    dni = (dzis - start).days
    tyg_uplynelo = dni / 7.0
    tyg_zrobione = godz_zrob / cap if cap else 0.0
    zapas_tyg = tyg_zrobione - max(tyg_uplynelo, 0.0)

    # Który sprint jest „bieżący": pierwszy z niedomkniętymi pozycjami.
    biezacy = next((s["n"] for s in sprinty if s["zrobione"] < s["pozycje"]), None)

    # Prognoza końca: tempo dotychczasowe albo — gdy plan jeszcze nie ruszył —
    # nominalna pojemność. Nie zgadujemy przyspieszenia, którego nie widać.
    pozostale_godz = godz_razem - godz_zrob
    tempo = cap  # h/tydzień
    tygodni_do_konca = pozostale_godz / tempo if tempo else 0
    koniec_nominalny = daty(D, max(D["sprinty"]))[1]
    koniec_prognoza = dzis + datetime.timedelta(weeks=tygodni_do_konca)

    blokery_otwarte = [
        r[0] for r in D["macierz"]
        if r[11].strip() == "BLOKER STARTU" and r[10] not in ZROBIONE_WERDYKTY
    ]

    return {
        "dzis": str(dzis),
        "start": str(start),
        "cap": cap,
        "sprinty": sprinty,
        "biezacySprint": biezacy,
        "godzinyRazem": godz_razem,
        "godzinyZrobione": godz_zrob,
        "pozycjeRazem": poz_razem,
        "pozycjeZrobione": poz_zrob,
        "procent": round(100.0 * godz_zrob / godz_razem, 1) if godz_razem else 0.0,
        "tygodnieUplynelo": round(tyg_uplynelo, 2),
        "tygodnieZrobione": round(tyg_zrobione, 2),
        "zapasTygodni": round(zapas_tyg, 2),
        "koniecNominalny": str(koniec_nominalny),
        "koniecPrognoza": str(koniec_prognoza),
        "blokeryOtwarte": blokery_otwarte,
    }


def buduj_dashboard_planu(D):
    H = D["cfg"]["godziny_nakladu"]
    EP = {e["id"]: e for e in D["epiki"]}
    R, _ = reszta(D)
    P = {
        "tasks": zadania(D),
        "sprints": [{"n": n, "cel": D["sprinty"][n]["cel"], "od": str(daty(D, n)[0]),
                     "do": str(daty(D, n)[1]), "faza": faza(D, n), "ryzyko": D["sprinty"][n]["ryzyko"]}
                    for n in sorted(D["sprinty"])],
        "fazy": [{"od": f_["od"], "do": f_["do"], "tytul": f_["tytul"], "opis": f_["opis"]} for f_ in D["fazy"]],
        "rest": [[r[0], epik_dla(D, r), EP[epik_dla(D, r)]["nazwa"], EP[epik_dla(D, r)]["kw"], r[2],
                  D["cfg"]["kategorie"].get(r[1], r[1]), r[11], H.get(r[12], 0)] for r in R],
        "epiki": [[e["id"], e["nazwa"], e["prio"], e["kw"], e["zakres"], e["why"]] for e in D["epiki"]],
        "oos": [[r[0], r[2], D["cfg"]["kategorie"].get(r[1], r[1]), r[13]]
                for r in D["macierz"] if r[10] == "POZA ZAKRESEM"],
        "cap": D["cfg"]["sprint_godzin"],
        "postep": postep(D),
    }
    tpl = (SZAB / "dashboard_plan.html").read_text(encoding="utf-8")
    OUT_P.mkdir(exist_ok=True)
    (OUT_P / "VERRIS_PLAN_DASHBOARD.html").write_text(
        tpl.replace("__DATA__", json.dumps(P, ensure_ascii=False)), encoding="utf-8")


# ───────────────────────────── 6. szkielety dokumentacji zadań ─────────────────────────────
def buduj_szkielety_zadan(D, n):
    """Tworzy docs/zadania/<ID>-<slug>.md dla sprintu n, wypełnione tym, co wiadomo z macierzy."""
    import unicodedata
    kat = REPO / "docs" / "zadania"
    kat.mkdir(parents=True, exist_ok=True)

    def slug(t):
        t = t.lower().replace("ł", "l").replace("ż", "z").replace("ź", "z")
        t = unicodedata.normalize("NFKD", t)
        t = "".join(c for c in t if not unicodedata.combining(c))
        t = re.sub(r"[^a-z0-9]+", "-", t).strip("-")
        return "-".join(t.split("-")[:5])

    d0, d1 = daty(D, n)
    utworzone, pominiete = [], []
    for i, tytul, h, typ, prio, dow, dod, ctx in pozycje_sprintu(D, n):
        sciezka = kat / f"{i}-{slug(tytul)}.md"
        if sciezka.exists():
            pominiete.append(sciezka.name)
            continue
        r = D["wg_id"].get(i)
        stan = r[8] if r else "—"
        dowod_blok = (f"```\n{dow}\n```\n" if dow not in ("—", "") else "_Pozycja spoza audytu — dowodu z kodu nie ma._\n")
        L = [
            f"# `{i}` — {tytul}\n",
            "| | |", "|---|---|",
            f"| **Sprint** | {n} ({d0} – {d1}) |",
            f"| **Priorytet** | {prio} |",
            f"| **Nakład** | planowany {h} h · rzeczywisty ? h |",
            f"| **Zależy od** | {D['cfg']['zaleznosci'].get(i, '—')} |",
            "| **Status** | do zrobienia |",
            "| **Data zamknięcia** | |\n",
            "---\n",
            "## Problem\n",
            (ctx if ctx.strip() else "_Do uzupełnienia._") + "\n",
            "## Dowód przed\n",
            dowod_blok,
            f"**Stan w macierzy przed:** `{stan}`\n",
            "## Rozwiązanie\n", "_Do uzupełnienia w trakcie pracy. Zapisz też podejścia odrzucone i dlaczego._\n",
            "## Zmienione pliki\n", "| Plik | Co się zmieniło |", "|---|---|", "| | |\n",
            "Migracje bazy: —  ", "Zmienne środowiskowe: —\n",
            "## Testy\n", "| Test | Co sprawdza |", "|---|---|", "| | |\n",
            "**Czy test najpierw czerwienił się na starym kodzie?** —\n",
            "## Dowód po\n", "_`plik:linia` wskazujące na implementację — to trafia do macierzy._\n",
            "**Osiągnięty poziom dowodu:**", "- [ ] D1 — kod istnieje", "- [ ] D2 — test przechodzi w CI",
            "- [ ] D3 — zaobserwowane na produkcji (data)", "- [ ] D4 — powtarzalna procedura z właścicielem i datą\n",
            "**Stan w macierzy po:** \n",
            "## Definicja ukończenia\n", f"> {dod}\n",
            "## Czego to nadal nie robi\n",
            "_Jeżeli lista nie jest pusta, stan w macierzy to `CZĘŚCIOWE`, a brakująca część wraca do backlogu z nowym ID._\n",
            "## Ryzyko i wycofanie\n", "_Co może pójść źle i jak cofnąć._\n",
            "## Wpływ na inne pozycje\n", "_Które ID z macierzy to zamyka, otwiera albo zmienia._\n",
        ]
        sciezka.write_text("\n".join(L), encoding="utf-8")
        utworzone.append(sciezka.name)
    return utworzone, pominiete


# ───────────────────────────── status ─────────────────────────────
def status(D):
    from collections import Counter
    R, RH = reszta(D)
    total = sum(godziny(D, n) for n in D["sprinty"])
    blk = [r for r in D["macierz"] if r[11] == "BLOKER STARTU"]
    print(f"Macierz: {len(D['macierz'])} pozycji")
    print("  werdykty:", dict(Counter(r[10] for r in D["macierz"])))
    print("  stany:   ", dict(Counter(r[8] for r in D["macierz"])))
    print(f"  BLOKERY STARTU: {len(blk)}")
    for r in blk:
        n = [k for k in D["sprinty"] if r[0] in {x[0] for x in pozycje_sprintu(D, k)}]
        print(f"    {r[0]:6} sprint {n[0] if n else '?':>2}  {r[2][:60]}")
    print(f"\nPlan: {len(D['sprinty'])} sprintów, {total} h "
          f"(pojemność {len(D['sprinty'])*D['cfg']['sprint_godzin']} h)")
    print(f"Po starcie: {len(R)} pozycji, {RH} h")


# ───────────────────────────── main ─────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Generator artefaktów audytu i planu Verris")
    ap.add_argument("--sprawdz", action="store_true", help="tylko walidacja spójności danych")
    ap.add_argument("--status", action="store_true", help="krótkie podsumowanie stanu")
    ap.add_argument("--zadania", type=int, metavar="SPRINT",
                    help="utwórz szkielety docs/zadania/ dla danego sprintu (nie nadpisuje istniejących)")
    a = ap.parse_args()

    D = wczytaj()
    bledy, ostrz = sprawdz(D)
    for b in bledy:
        print("BŁĄD:      ", b, file=sys.stderr)
    for o in ostrz:
        print("OSTRZEŻENIE:", o, file=sys.stderr)
    if bledy:
        print(f"\n{len(bledy)} błędów — nie generuję.", file=sys.stderr)
        return 1
    if a.sprawdz:
        print(f"Dane spójne. Ostrzeżeń: {len(ostrz)}.")
        return 0
    if a.status:
        status(D)
        return 0
    if a.zadania:
        if a.zadania not in D["sprinty"]:
            print(f"Nie ma sprintu {a.zadania}.", file=sys.stderr)
            return 1
        nowe, stare = buduj_szkielety_zadan(D, a.zadania)
        for f_ in nowe:
            print("utworzono  docs/zadania/" + f_)
        for f_ in stare:
            print("pominięto  docs/zadania/" + f_ + " (już istnieje)")
        return 0

    buduj_macierz_xlsx(D)
    buduj_dashboard_luk(D)
    buduj_backlog_xlsx(D)
    buduj_plan_md(D)
    buduj_dashboard_planu(D)
    print("Zbudowane:")
    for p in [OUT_A / "VERRIS_PARYTET_FUNKCJI_2026-08.xlsx", OUT_A / "VERRIS_LUKI_DASHBOARD.html",
              OUT_P / "VERRIS_BACKLOG_STARTOWY.xlsx", OUT_P / "PLAN_SPRINTOW_2026-08.md",
              OUT_P / "VERRIS_PLAN_DASHBOARD.html"]:
        print("  ", p.relative_to(REPO))
    print("\nUWAGA: formuły w XLSX nie mają zapisanych wartości, dopóki plik nie zostanie otwarty "
          "w arkuszu kalkulacyjnym. To normalne — Excel i LibreOffice przeliczą je przy otwarciu.")
    if ostrz:
        print(f"\nOstrzeżeń do przejrzenia: {len(ostrz)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
